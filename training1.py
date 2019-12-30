import openpyxl
wb = openpyxl.load_workbook('example.xlsx')

# 1. Getting SheetName from the workbook

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

# 2. Creating Sheets from the workbook
mysheet = wb.create_chartsheet('mySheet')
print(wb.sheetnames)

# 3. Getting Sheets from the workbook
#sheet3 = wb.get_sheet_by_name('Sheet3')
#sheet4 = wb['mySheet']

#4. Getting cells from the sheets

# 4.1
ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)

# 4.2 此方法适用于for 循环
print(ws.cell(row=1, column=2))
print(ws.cell(row=1, column=2).value)

for i in range(2,11):
    print(ws.cell(row=i, column=1).value,ws.cell(row=i, column=2).value)

# 4.3 getting rows and columns for the sheet
colC = ws['C']
print(colC) #得到列序
print(colC[2].value) #打印 C2 单元格

row6 = ws['6']
col_range = ws['B:C']
row_range = ws[2:6]

# 4.4 列切片：按列打印
for col in col_range:
    for cell in col:
        print(cell.value)
# 4.5 行切片：按行打印

for row in row_range:
    for cell in row:
        print(cell.value)
# 4.6 方法函数，生成器：按行打印
for row in ws.iter_rows(min_row=1, max_row=3, max_col=2):
    print(cell)

#4.7 运用元组
print(tuple(ws.rows))        #由内而外： 内部是以行cell组成的元组，外部是以行组成的元组

#4.8 以单元格为对象为单元格
cell_range = ws['A2:C6']
for RowOfboject in cell_range:
    for cellObj in RowOfboject:
        print(cellObj.coordinate,cellObj.value)
# 4.9 reference coordinate
# 4.9.1
c = ws['B1']
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))

# 4.9.2
print('Cell {} is {}\n'.format(c.coordinate, c.value))

# 5 获得列的数字形式
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(2),get_column_letter(20))
print(column_index_from_string('B'),column_index_from_string('Z'))

# 打印sheet的单元格最大范围
print('{} * {}'.format(ws.max_row,ws.max_column))