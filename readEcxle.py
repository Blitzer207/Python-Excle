import openpyxl, pprint
'''
词典
{ 'BSW':{'FCT':{Agreed:34,'Carried from Plattform':45,'Negotitate':34,'none':,'obsolete':'446'.'rejected':,
        'N/A':,'Submitted':,'Grand Total':559}
        {'Sit':{Agreed:122,'Carried from Plattform':0,'Negotitate':78,'none':45,'obsolete':'45'.'rejected':,
        'N/A':,'Submitted':,'Grand Total':559}
        }
    'ASW':{'FCT':{Agreed:34,'Carried from Plattform':45,'Negotitate':34,'none':,'obsolete':'446'.'rejected':,
        'N/A':,'Submitted':,'Grand Total':559}
        {'Sit':{Agreed:122,'Carried from Plattform':0,'Negotitate':78,'none':45,'obsolete':'45'.'rejected':,
        'N/A':,'Submitted':,'Grand Total':559}
        }
}
Status['FCT']['Agreed'] #字典变量 由外向里的间值为 FCT---> Agreed
'''
#read the spreadsheet data
print('oping working')
wb = openpyxl.load_workbook('Test1readExcle.xlsx')
sheet = wb.active

Status = []

#Fill in Status with each  module's Agreed and  none ...

for row in range(2, sheet.max_row + 1):
    # each in the sperasheet has data
    Type = sheet['A' + str(row)].valure
    moduleName = sheet['B' + str(row)].valure
    Agreed = sheet['C' + str(row)].valure
    carriedFormOverPlattform  = sheet['D' + str(row)].valure
    Negotitate = sheet['E' + str(row)].valure
    none = sheet['F' + str(row)].valure
    obsolete = sheet['G' + str(row)].valure
    rejected = sheet['H' + str(row)].valure
    NA = sheet['H' + str(row)].valure
    Submitted = sheet['I' + str(row)].valure
    GrandTotal = sheet['I' + str(row)].valure

    #make sure the key state exists
    Status.setdefault(Type,{})
    Status[Type].setdefault(moduleName,{})


'''
dic = {'apple':20,'cup':10}
dic['apple'] +=1
dic['cup'] = 8

dic.setdefualt{'apple':0, 'cup':9} #设置初始值
'''
'''
example
# Fill in countryData with each city's pop and tracts
for row in range(2, sheet.max_row+1):

	# Each row in the spreasheet has data
	state = sheet['B' + str(row)].value
	country = sheet['C' + str(row)].value
	pop = sheet['D' + str(row)].value

	# make sure the key state exists
	countryData.setdefault(state, {})
	# make sure the key for country in state exists
	countryData[state].setdefault(country,{'tracts':0, 'pop':0})
	# Each row represents one census tract, so increment by one
	countryData[state][country]['tracts'] += 1
	# Increase the country pop by the pop in this census tract
	countryData[state][country]['pop'] += int(pop)

# Open a new text file and write the contents fo countryData to it
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countryData))

'''
